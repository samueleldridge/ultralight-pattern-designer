"""
Data export module.
CSV, Excel, and PDF export functionality.
"""

import csv
import io
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
from dataclasses import dataclass


@dataclass
class ExportOptions:
    """Options for data export"""
    filename: Optional[str] = None
    include_headers: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    max_rows: int = 10000
    sheet_name: str = "Data"


class CSVExporter:
    """Export data to CSV format"""
    
    @staticmethod
    def export(
        data: List[Dict[str, Any]],
        options: Optional[ExportOptions] = None
    ) -> bytes:
        """
        Export data to CSV.
        Returns bytes for HTTP response.
        """
        options = options or ExportOptions()
        
        if not data:
            return b""
        
        # Limit rows
        data = data[:options.max_rows]
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        if options.include_headers:
            headers = list(data[0].keys())
            writer.writerow(headers)
        
        # Write data
        for row in data:
            formatted_row = []
            for value in row.values():
                formatted_row.append(CSVExporter._format_value(value, options))
            writer.writerow(formatted_row)
        
        return output.getvalue().encode('utf-8')
    
    @staticmethod
    def _format_value(value: Any, options: ExportOptions) -> str:
        """Format a value for CSV export"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime(options.date_format)
        if isinstance(value, (dict, list)):
            import json
            return json.dumps(value)
        return str(value)
    
    @staticmethod
    def get_content_type() -> str:
        return "text/csv; charset=utf-8"
    
    @staticmethod
    def get_filename(base_name: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_{timestamp}.csv"


class ExcelExporter:
    """Export data to Excel format"""
    
    def __init__(self):
        self._has_openpyxl = False
        try:
            import openpyxl
            self._has_openpyxl = True
        except ImportError:
            pass
    
    def export(
        self,
        data: List[Dict[str, Any]],
        options: Optional[ExportOptions] = None
    ) -> bytes:
        """
        Export data to Excel (.xlsx).
        Returns bytes for HTTP response.
        """
        if not self._has_openpyxl:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        options = options or ExportOptions()
        
        if not data:
            return b""
        
        # Limit rows
        data = data[:options.max_rows]
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = options.sheet_name
        
        # Write headers
        if options.include_headers:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        start_row = 2 if options.include_headers else 1
        for row_idx, row in enumerate(data, start_row):
            for col_idx, value in enumerate(row.values(), 1):
                formatted_value = self._format_value(value, options)
                ws.cell(row=row_idx, column=col_idx, value=formatted_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def _format_value(value: Any, options: ExportOptions) -> Any:
        """Format a value for Excel export"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value
        if isinstance(value, (dict, list)):
            import json
            return json.dumps(value)
        return value
    
    @staticmethod
    def get_content_type() -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    @staticmethod
    def get_filename(base_name: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_{timestamp}.xlsx"


class PDFExporter:
    """Export data to PDF format"""
    
    def __init__(self):
        self._has_reportlab = False
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            self._has_reportlab = True
        except ImportError:
            pass
    
    def export(
        self,
        data: List[Dict[str, Any]],
        title: str = "Report",
        options: Optional[ExportOptions] = None
    ) -> bytes:
        """
        Export data to PDF.
        Returns bytes for HTTP response.
        """
        if not self._has_reportlab:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        
        options = options or ExportOptions()
        
        if not data:
            return b""
        
        # Limit rows
        data = data[:options.max_rows]
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 0.25*inch))
        
        # Generated timestamp
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime(options.date_format)}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.25*inch))
        
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data:
            formatted_row = []
            for value in row.values():
                formatted_value = self._format_value(value, options)
                # Limit cell content length
                if len(str(formatted_value)) > 100:
                    formatted_value = str(formatted_value)[:97] + "..."
                formatted_row.append(formatted_value)
            table_data.append(formatted_row)
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ])
        table.setStyle(style)
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def _format_value(value: Any, options: ExportOptions) -> str:
        """Format a value for PDF export"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime(options.date_format)
        if isinstance(value, (dict, list)):
            import json
            return json.dumps(value)[:100]
        return str(value)
    
    @staticmethod
    def get_content_type() -> str:
        return "application/pdf"
    
    @staticmethod
    def get_filename(base_name: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_{timestamp}.pdf"


class ExportManager:
    """
    Manage data exports in multiple formats.
    """
    
    FORMATS = {
        'csv': CSVExporter,
        'excel': ExcelExporter,
        'pdf': PDFExporter
    }
    
    def __init__(self):
        self._exporters = {
            'csv': CSVExporter(),
            'excel': ExcelExporter(),
            'pdf': PDFExporter()
        }
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        format: str,
        filename_base: str = "export",
        options: Optional[ExportOptions] = None
    ) -> Dict[str, Any]:
        """
        Export data in specified format.
        Returns dict with content bytes and metadata.
        """
        if format not in self.FORMATS:
            raise ValueError(f"Unsupported format: {format}. Use: {list(self.FORMATS.keys())}")
        
        options = options or ExportOptions()
        exporter = self._exporters[format]
        
        # Generate export
        if format == 'csv':
            content = exporter.export(data, options)
        elif format == 'excel':
            content = exporter.export(data, options)
        elif format == 'pdf':
            content = exporter.export(data, filename_base, options)
        else:
            raise ValueError(f"Unknown format: {format}")
        
        # Get filename
        filename_method = getattr(self.FORMATS[format], 'get_filename')
        filename = filename_method(filename_base)
        
        # Get content type
        content_type_method = getattr(self.FORMATS[format], 'get_content_type')
        content_type = content_type_method()
        
        return {
            "content": content,
            "filename": filename,
            "content_type": content_type,
            "format": format,
            "row_count": len(data[:options.max_rows]),
            "size_bytes": len(content)
        }
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        formats = ['csv']  # CSV always supported
        
        # Check optional dependencies
        try:
            import openpyxl
            formats.append('excel')
        except ImportError:
            pass
        
        try:
            import reportlab
            formats.append('pdf')
        except ImportError:
            pass
        
        return formats


class ScheduledExport:
    """
    Scheduled export configuration.
    """
    
    def __init__(
        self,
        name: str,
        query: str,
        format: str,
        schedule: str,  # cron expression
        recipients: List[str],
        options: Optional[ExportOptions] = None
    ):
        self.name = name
        self.query = query
        self.format = format
        self.schedule = schedule
        self.recipients = recipients
        self.options = options or ExportOptions()
        self.last_run: Optional[datetime] = None
        self.last_status: Optional[str] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "query": self.query,
            "format": self.format,
            "schedule": self.schedule,
            "recipients": self.recipients,
            "last_run": self.last_run.isoformat() if self.last_run else None,
            "last_status": self.last_status
        }
