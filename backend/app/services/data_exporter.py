"""
Data Export Service

Export query results in multiple formats for sharing with stakeholders:
- CSV (spreadsheet-compatible)
- Excel (formatted with charts)
- PDF (presentation-ready)
- JSON (for technical users)

Features:
- Async processing for large datasets
- Progress tracking
- Email delivery
- Custom formatting for presentations
"""

import csv
import io
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
from dataclasses import dataclass
from enum import Enum


class ExportFormat(str, Enum):
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"
    JSON = "json"


class ExportStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETE = "complete"
    FAILED = "failed"


@dataclass
class ExportJob:
    id: str
    user_id: str
    tenant_id: str
    query: str
    format: ExportFormat
    status: ExportStatus
    created_at: datetime
    completed_at: Optional[datetime] = None
    file_url: Optional[str] = None
    file_size: Optional[int] = None
    row_count: Optional[int] = None
    error_message: Optional[str] = None


class DataExporter:
    """Export query results in various formats"""
    
    def __init__(self, storage_service=None):
        self.storage = storage_service
    
    async def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> bytes:
        """Export data to CSV format"""
        if not data:
            return b""
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = list(data[0].keys())
        writer.writerow(headers)
        
        # Write data rows
        for row in data:
            writer.writerow([self._format_value(row.get(h)) for h in headers])
        
        return output.getvalue().encode('utf-8')
    
    async def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        title: str = "Export",
        description: str = ""
    ) -> bytes:
        """Export data to Excel format with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        if not data:
            return b""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Title row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Description row
        if description:
            ws.merge_cells('A2:E2')
            desc_cell = ws['A2']
            desc_cell.value = description
            desc_cell.font = Font(size=10, italic=True)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 20
            start_row = 3
        else:
            start_row = 2
        
        # Headers
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        header_font = Font(bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=self._format_excel_value(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in data[:100]:  # Sample first 100 rows
                cell_value = str(row.get(header, ""))
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Summary row
        summary_row = start_row + len(data) + 2
        ws.merge_cells(f'A{summary_row}:E{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = f"Exported {len(data)} rows on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        summary_cell.font = Font(size=9, italic=True, color="666666")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    async def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str = "Data Export",
        query: str = ""
    ) -> bytes:
        """Export data to PDF format (presentation-ready)"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            raise ImportError("reportlab not installed. Run: pip install reportlab")
        
        if not data:
            return b""
        
        output = io.BytesIO()
        
        # Use landscape for wide tables
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#6366F1'),
            spaceAfter=12
        )
        elements.append(Paragraph(title, title_style))
        
        # Query description
        if query:
            query_style = ParagraphStyle(
                'QueryStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                fontName='Courier',
                spaceAfter=12
            )
            elements.append(Paragraph(f"Query: {query[:100]}...", query_style))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20
        )
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data[:1000]:  # Limit to 1000 rows for PDF
            table_data.append([self._format_pdf_value(row.get(h)) for h in headers])
        
        if len(data) > 1000:
            elements.append(Paragraph(f"Showing first 1000 of {len(data)} rows", styles['Normal']))
            elements.append(Spacer(1, 12))
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1  # Center
        )
        elements.append(Paragraph("Generated by AI Analytics Platform", footer_style))
        
        doc.build(elements)
        output.seek(0)
        return output.getvalue()
    
    async def export_to_json(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to JSON format"""
        return json.dumps(data, indent=2, default=str).encode('utf-8')
    
    def _format_value(self, value: Any) -> str:
        """Format a value for CSV export"""
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return str(value)
        if isinstance(value, datetime):
            return value.isoformat()
        return str(value)
    
    def _format_excel_value(self, value: Any) -> Any:
        """Format a value for Excel export"""
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)  # Excel doesn't like timezone-aware dates
        return value
    
    def _format_pdf_value(self, value: Any) -> str:
        """Format a value for PDF export"""
        if value is None:
            return "-"
        if isinstance(value, float):
            return f"{value:,.2f}"
        if isinstance(value, int):
            return f"{value:,}"
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        return str(value)[:50]  # Truncate long strings
    
    async def export(
        self,
        data: List[Dict[str, Any]],
        format: ExportFormat,
        title: str = "Data Export",
        query: str = ""
    ) -> bytes:
        """Export data in specified format"""
        if format == ExportFormat.CSV:
            return await self.export_to_csv(data)
        elif format == ExportFormat.EXCEL:
            return await self.export_to_excel(data, title, query)
        elif format == ExportFormat.PDF:
            return await self.export_to_pdf(data, title, query)
        elif format == ExportFormat.JSON:
            return await self.export_to_json(data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def get_content_type(self, format: ExportFormat) -> str:
        """Get HTTP content type for format"""
        types = {
            ExportFormat.CSV: "text/csv",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.PDF: "application/pdf",
            ExportFormat.JSON: "application/json"
        }
        return types.get(format, "application/octet-stream")
    
    def get_extension(self, format: ExportFormat) -> str:
        """Get file extension for format"""
        extensions = {
            ExportFormat.CSV: ".csv",
            ExportFormat.EXCEL: ".xlsx",
            ExportFormat.PDF: ".pdf",
            ExportFormat.JSON: ".json"
        }
        return extensions.get(format, ".dat")
